import openpyxl
import pprint
import logging
wb = openpyxl.load_workbook('Test_Result.xlsx')
# print all sheet in excel
print(wb.sheetnames)

# read date in sheet choose in excel
sheet = wb['LTPv20210927_BSP5.10_Unify']
cellA6 = sheet['A6']
print(cellA6.value)

cellB6 = sheet['B6']
print(cellB6.value)
cell = sheet.cell(row=6, column=4)
print(cell.value)

cells_tuple = sheet['A6:C10']
pprint.pprint(cells_tuple)
# lay gia tri cua 1 cell
print(cells_tuple[0][1].value)

# lay toan bo cac o trong pham vi duoi dang 1 list
g = sheet.iter_rows(min_row=2, max_row=6, min_col=2, max_col=3)
print(type(g))
# <class 'generator'
cells_list=list(g)
pprint.pprint(cells_list)
print(cells_list[0][1].value)

# list all data from sheet excel
g_all = sheet.values
pprint.prrint(list(g_all), width=40)
 def get_cell_value_list(sheet):
     return([[cell.value for cell in row] for row in sheet])
 all_cell_value = get_cell_value_list(sheet)
 pprint.prrint(all_cell_value, width=40)
 








